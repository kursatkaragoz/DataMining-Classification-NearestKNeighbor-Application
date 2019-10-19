from tkinter import *
from tkinter import Tk, Label, Button
from tkinter import ttk
from tkinter import filedialog
import openpyxl
import pymysql.cursors
from tkinter import messagebox
import math
import matplotlib.pyplot as plt
from matplotlib.figure import Figure
from numpy import arange, sin, pi
import numpy as np
import pandas as pd
class görsel():
    def __init__(self,pencere):
        self.pencere = pencere
        pencere.title("K EN YAKIN KOMŞU ALGORİTMASI")
        pencere.geometry("750x450+300+120")
        self.scrollbar = Scrollbar(pencere)
        self.listekutusu = Listbox(pencere , width = 48 , height = 27 , yscrollcommand = self.scrollbar.set)
        self.listekutusu.place(x=20 , y = 0)
        self.scrollbar.pack(side=LEFT , fill=Y)
        self.scrollbar.config(command=self.listekutusu.yview)


        self.dosyasecButton = Button(pencere,text = "Dosya Seciniz" ,width=12 , height = 2,command = self.dosyaSec)
        self.dosyasecButton.place(x=610,y=80)
        self.grafikButton = Button(pencere, text="Degerleri Göser", width=12, height=2, command=self.GrafikGöster)
        self.grafikButton.place(x=610, y=120)
        self.uzaklıkButton = Button(pencere, text="UzaklıkHesapla", width=12, height=2, command=self.Uzaklik_Hesapla)
        self.uzaklıkButton.place(x=610, y=160)
        self.uzaklik_grafik = Button(pencere,text = "Uzaklik Grafikk" ,width=12 , height = 2 , command = self.Uzaklik_Grafik_Göster)
        self.uzaklik_grafik.place(x=610,y=200)
        self.temizleButton = Button(pencere, text="Temizle", width=12, height=2, command=self.Sifirla)
        self.temizleButton.place(x=610, y=240)

        self.kdegerLabel = Label(pencere ,text="K Degeri Giriniz :")
        self.kdegerLabel.place(x=320,y=10)
        self.kdeger = ttk.Combobox(pencere   , width=6 , state="disabled" )
        self.kdeger.bind('<<ComboboxSelected>>',self.k_deger)

        self.kdeger.place(x=470,y=10)
        self.uzaklıkNoktalariLabel = Label(pencere , text="X1 X2 X3 Noktalarını Giriniz :")

        self.uzaklıkNoktalariLabel.place(x=320 , y=40)
        self.nokta1 = ttk.Combobox(pencere , state='disabled' ,textvariable ="nokta1" , width=6 , )
        self.nokta1.place(x=470,y=40)
        self.nokta2 = ttk.Combobox(pencere , state='disabled' ,textvariable ="nokta2" , width=6 , )
        self.nokta2.place(x=530,y=40)
        self.nokta3 = ttk.Combobox(pencere , state='disabled' ,textvariable ="nokta3" , width=6 , )
        self.nokta3.place(x=590,y=40)
        self.CheckVar1 = IntVar()
        self.normalize = Checkbutton(pencere ,state="disabled", text="Normalizasyon Değerlerinden Yararlan" , variable = self.CheckVar1 , onvalue = 1 , offvalue = 0 ,height = 1, width=30 , command = self.checkbutton)
        self.normalize.place(x=320,y=70)
        self.CheckVar2 = IntVar()
        self.k_algoritma = Checkbutton(pencere , text="K En Küçük" , variable = self.CheckVar2, onvalue = 1 , offvalue = 0 ,height = 1, width=15 )#, command = self.checkbutton)
        self.k_algoritma.place(x=310,y=250)
        self.CheckVar3 = IntVar()
        self.agirlik_algoritma = Checkbutton(pencere , text="Ağırlık Oylama" , variable = self.CheckVar3 , onvalue = 1 , offvalue = 0 ,height = 1, width=15)# , command = self.checkbutton)
        self.agirlik_algoritma.place(x=420,y=250)
        self.SinifiBul = Button(pencere,text = "Sinifi Tahmin Et" ,width=24 , height = 2,command = self.Sinifi_Tahminet)
        self.SinifiBul.place(x=345,y=280)
        self.degerler = []
        for i in range(2,32):
           self.degerler.append(i)
        self.kdeger["values"]= self.degerler
        self.nokta1["values"]= self.degerler
        self.nokta2["values"]= self.degerler
        self.nokta3["values"]= self.degerler


        self.x1min =Label(pencere ,text="X1 min: ")
        self.x1min.place(x=320,y=120)
        self.x1minkutu=Entry(pencere , width=5 , state="readonly")
        self.x1minkutu.place(x=370,y=120)
        self.x1max =Label(pencere ,text="X1 max: ")
        self.x1max.place(x=420,y=120)
        self.x1maxkutu=Entry(pencere , width=5 , state="readonly")
        self.x1maxkutu.place(x=470,y=120)

        self.x2min =Label(pencere ,text="X2 min: ")
        self.x2min.place(x=320,y=150)
        self.x2minkutu=Entry(pencere , width=5 , state="readonly")
        self.x2minkutu.place(x=370,y=150)
        self.x2max =Label(pencere ,text="X2 max: ")
        self.x2max.place(x=420,y=150)
        self.x2maxkutu=Entry(pencere , width=5 , state="readonly")
        self.x2maxkutu.place(x=470,y=150)

        self.x3min =Label(pencere ,text="X3 min: ")
        self.x3min.place(x=320,y=180)
        self.x3minkutu=Entry(pencere , width=5 , state="readonly")
        self.x3minkutu.place(x=370,y=180)
        self.x3max =Label(pencere ,text="X3 max: ")
        self.x3max.place(x=420,y=180)
        self.x3maxkutu=Entry(pencere , width=5 , state="readonly")
        self.x3maxkutu.place(x=470,y=180)

        self.yeninokta1text=Label(pencere ,text="X1'': ")
        self.yeninokta1text.place(x=330,y=210)
        self.yeninokta1=Entry(pencere , width=5 , state="readonly")
        self.yeninokta1.place(x=370,y=210)
        self.yeninokta2text =Label(pencere ,text="X2'': ")
        self.yeninokta2text.place(x=430,y=210)
        self.yeninokta2=Entry(pencere , width=5 , state="readonly")
        self.yeninokta2.place(x=470,y=210)
        self.yeninokta3text =Label(pencere ,text="X3'': ")
        self.yeninokta3text.place(x=520,y=210)
        self.yeninokta3=Entry(pencere , width=5 , state="readonly")
        self.yeninokta3.place(x=560,y=210)

        self.Sifirla()

    def k_deger(self,event=None):
        self.listekutusu.insert(END,"")
        self.listekutusu.insert(END,"K degeri {} olarak belirlendi".format(self.kdeger.get()))
        self.normalize.configure(state="active")


    def Sinifi_Tahminet(self):

        self.k_uzaklik = []
        self.k_durum = []
        self.x= self.kdeger.get()
        kötü = "KÖTÜ"
        iyi = "İYİ"
        yenisinif = ""
        kotu=0;
        ıyı=0;
        if(self.CheckVar2.get() == 1):
            if(self.kdeger.get()==""):
                messagebox.showinfo("UNKOWN VALUE","K Değerini Belirleyiniz")
            else:
                    connection = pymysql.connect(host='localhost',
                                             user='root',
                                             password='',
                                             db='verimadenciligi',
                                             charset='utf8mb4',
                                             cursorclass=pymysql.cursors.DictCursor)
                    with connection.cursor() as cursor:
                        sql = "select  uzaklık_deger , durum  FROM `degerler` ORDER BY uzaklık_deger LIMIT %s" % self.x
                        cursor.execute(sql)
                        for abc in cursor.fetchall():
                            self.k_uzaklik.append(abc["uzaklık_deger"])
                            self.k_durum.append(abc["durum"])

                    self.listekutusu.insert(END,"------------------------")
                    self.listekutusu.insert(END,"------------------------")
                    self.listekutusu.insert(END,"En Yakın K Komşu ile Tahmin:".format(self.x))
                    self.listekutusu.insert(END,"İlk {} eleman uzaklıkları ve durumları".format(self.x))
                    for i in range(0,len(self.k_uzaklik)):
                        self.listekutusu.insert(END,"{:^10}{:^10}".format(self.k_uzaklik[i],self.k_durum[i]))
                        if(self.k_durum[i] ==kötü):
                            kotu=kotu+1
                        else:
                            ıyı = ıyı + 1
                    self.listekutusu.insert(END,"{} Adet iyi sınıf değeri ".format(ıyı))
                    self.listekutusu.insert(END,"{} Adet iyi sınıf değeri ".format(kotu))
                    if(kotu>ıyı):
                        yenisinif = "KÖTÜ"
                        self.listekutusu.insert(END,"Kötü Durum Sayısı > İyi Durum Sayısı")
                    elif((ıyı>kotu)):
                        yenisinif= "İYİ"
                        self.listekutusu.insert(END,"İyi Durum Sayısı > Kötü Durum Sayısı")
                    else:
                        yenisinif = "Eşit Olasılık bu durumda ağırlık oynama yönteminden gelecek sonuca bakılır"
                    self.listekutusu.insert(END,"Referans noktalarının tahmini sınıfı : {}".format(yenisinif))
                    messagebox.showinfo("BAŞARI","Tahmin Etme İşlemi Gerçekleşti")



        ############################## AGİRLİK YÖNTEMİYLE SINIF BELİRLENİYOR
        if(self.CheckVar3.get() == 1):
            if(self.kdeger.get()==""):
                if(self.CheckVar2.get()==0):
                    messagebox.showinfo("UNKOWN VALUE","K Değerini Belirleyiniz")
            else:
                self.listekutusu.insert(END,"------------------------")
                self.listekutusu.insert(END,"------------------------")
                self.listekutusu.insert(END,"Ağırlık Oylama İle Tahmin Başladı:".format(self.x))
                self.agirlik_oylama()

        if((self.CheckVar3.get() == 0) & (self.CheckVar2.get()==0)):
            messagebox.showwarning("Class Unkown","Tahmin Yöntemi Seçiniz")



    def agirlik_oylama(self):

        connection = pymysql.connect(host='localhost',
                                         user='root',
                                         password='',
                                         db='verimadenciligi',
                                         charset='utf8mb4',
                                         cursorclass=pymysql.cursors.DictCursor)

        kötü = "KÖTÜ"
        iyi = "İYİ"
        yenisinif = ""
        kotu=0.0;
        ıyı=0.0;
        with connection.cursor() as cursor:
            self.uzaklik_deger = []
            self.islem_id = []
            self.uzaklik_durum = []
            self.uzaklik_deger= []
            self.agirlik_deger = []
            self.uzaklik_x1=[]
            self.uzaklik_x2=[]
            self.uzaklik_x3=[]
            sql = "SELECT islem_id,x1,x2,x3,uzaklık_deger,durum  FROM `degerler`"
            cursor.execute(sql)
            for abc in cursor.fetchall():
                self.uzaklik_deger.append(abc["uzaklık_deger"])
                self.uzaklik_durum.append(abc["durum"])
                self.islem_id.append(int(abc["islem_id"]))
                self.uzaklik_x1.append(abc["x1"])
                self.uzaklik_x2.append(abc["x2"])
                self.uzaklik_x3.append(abc["x3"])



        for i in range(0,len(self.uzaklik_deger)):
            print(self.uzaklik_deger[i])
            toplam= "{:.2f}".format(1.0 / (self.uzaklik_deger[i]*self.uzaklik_deger[i]))
           # toplam =" {:.2f}".format(toplam)
            self.agirlik_deger.append(toplam)
        self.listekutusu.insert(END,"Agirlik Değerleri :")
        if(self.sütun_sayisi == 4):
                self.listekutusu.insert(END,"{:^10}{:^10}{:^10}{:^10}{:^10}{:^10}".format("X1","X2","X3","Uzaklık'","Agirlik'","Durum'"))
        else:
                self.listekutusu.insert(END,"{:^10}{:^10}{:^10}{:^10}{:^10}".format("X1","X2","Uzaklik","Agirlik","Durum"))
        for i in range (0,len(self.islem_id)):
         if(self.sütun_sayisi == 4):

            self.listekutusu.insert(END,"{:^11}{:^11}{:^11}{:^1}{:^11}{:^11}".format(self.uzaklik_x1[i],self.uzaklik_x2[i],self.uzaklik_x3[i] , self.uzaklik_deger[i],self.agirlik_deger[i],self.uzaklik_durum[i]))


         else:
            self.listekutusu.insert(END,"{:^10}{:^10}{:^10}{:^10}{:^10}".format(self.uzaklik_x1[i],self.uzaklik_x2[i], self.uzaklik_deger[i],self.agirlik_deger[i],self.uzaklik_durum[i]))

        try:
            with connection.cursor() as cursor:
                for i in range(0,len(self.islem_id)):
                   cursor.execute("UPDATE degerler SET agırlık_deger='%f'  WHERE islem_id='%d'"%(float(self.agirlik_deger[i]),int(self.islem_id[i])))
        except:
            messagebox.showerror("Hata","Agirlik Deger Veri Tabanına eklenemedi")

        if(self.kdeger.get()==""):
                messagebox.showinfo("UNKOWN VALUE","K Değerini Belirleyiniz")
        else:

                self.distance =[]
                self.distance_status=[]
                self.weight=[]
                kotu =0
                ıyı=0
                with connection.cursor() as cursor:
                        sql = "select  uzaklık_deger , agırlık_deger ,durum  FROM `degerler` ORDER BY uzaklık_deger LIMIT %s" % self.x
                        cursor.execute(sql)
                        for abc in cursor.fetchall():
                            self.distance.append(abc["uzaklık_deger"])
                            self.distance_status.append(abc["durum"])
                            self.weight.append(abc["agırlık_deger"])

                self.listekutusu.insert(END,"------------------------")
                self.listekutusu.insert(END,"------------------------")
                self.listekutusu.insert(END,"Agirlik ile Tahmin:")
                self.listekutusu.insert(END,"İlk {} eleman agirlik ve durumları".format(self.x))

                for i in range(0,len(self.distance)):
                        self.listekutusu.insert(END,"{:^10}{:^10}{:^10}".format(self.distance[i],self.weight[i],self.distance_status[i]))
                        if(self.distance_status[i] ==kötü):
                            kotu=kotu+self.weight[i]
                        else:
                            ıyı = ıyı + self.weight[i]

                self.listekutusu.insert(END,"Kötü Durumun Agirlik Toplamı = {}".format(kotu))
                self.listekutusu.insert(END,"İyi Durumun Agirlik Toplamı = {}".format(ıyı))
                if(kotu>ıyı):
                        yenisinif = "KÖTÜ"
                        self.listekutusu.insert(END,"Kötü Durum,İyi durumdan Büyüktür.Tahmini sınıf:Kötü")
                elif((ıyı>kotu)):
                        yenisinif= "İYİ"
                        self.listekutusu.insert(END,"İyi Durum,Kötü durumdan büyüktür.Tahmini Sınıf:İyi")
                else:
                    yenisinif = "Eşit Olasılık bu durumda ağırlık oynama yönteminden gelecek sonuca bakılır"





    def Uzaklik_Grafik_Göster(self):
        try:
            connection = pymysql.connect(host='localhost',
                                             user='root',
                                             password='',
                                             db='verimadenciligi',
                                             charset='utf8mb4',
                                             cursorclass=pymysql.cursors.DictCursor)
            wb = openpyxl.load_workbook(self.filename)
            tablo = wb.get_sheet_by_name('Sheet1')
            tablo = wb.active
            sütun_sayisi = tablo.max_column
            self.x11=[]
            self.x22=[]
            self.x33=[]
            self.xuzaklik=[]
            self.durum = []
            self.Durum=[]
            self.sayilar = []
            with connection.cursor() as cursor:
                sql = "SELECT x1 , x2 , x3 , durum ,  uzaklık_deger  FROM `degerler`"
                cursor.execute(sql)
                for abc in cursor.fetchall():
                    self.x11.append(abc["x1"])
                    self.x22.append(abc["x2"])
                    self.xuzaklik.append(abc["uzaklık_deger"])
                    self.durum.append(abc["durum"])
                    if(sütun_sayisi == 4):
                        self.x33.append(abc["x3"])
                for i in range(0,len(self.durum)):
                    self.sayilar.append(i)
                    if(self.durum[i] == "KÖTÜ"):
                        self.Durum.append(1)
                    else:
                        self.Durum.append(2)

            y_post = np.arange(len(self.sayilar))
            plt.subplot(3, 1, 1)
            plt.title("Grafikler ")
            plt.xlabel("x1")
            plt.ylabel("x2")
            plt.bar(y_post,self.xuzaklik,align='center',alpha=0.5)
            plt.xticks(y_post,self.sayilar)
            plt.subplot(3, 1, 2)
            plt.title("Grafikler ")
            plt.xlabel("x1")
            plt.ylabel("x2")
            plt.scatter(self.x11,self.x22,s=self.xuzaklik)
            plt.grid(True)
            plt.subplot(3, 1, 3)
            plt.xlabel("x1")
            plt.ylabel("x2")
            bc2=plt.scatter(self.x11,self.x22,s=300,c=self.Durum)

            plt.grid(True)
            plt.show()
        except:
            messagebox.showerror("HATA","Grafikleri görebilmek için prosedüre uygun yol izleyiniz")



    def checkbutton (self):
       # try:
            if(self.CheckVar1.get() == 1):
                self.GozlemDegerleriniGoster()


        #except:
         #   print(traceback.format_exc())
          #  messagebox.showerror("HATA","Hesaplamak için Gerekli prosedürleri yapınız")
           # self.CheckVar1.set(0)

    def GozlemDegerleriniGoster (self):
        wb = openpyxl.load_workbook(self.filename)
        tablo = wb.get_sheet_by_name('Sheet1')
        tablo = wb.active
        self.sütünn_sayisi = tablo.max_column
        print(self.sütünn_sayisi)
        connection = pymysql.connect(host='localhost',
                                     user='root',
                                     password='',
                                     db='verimadenciligi',
                                     charset='utf8mb4',
                                     cursorclass=pymysql.cursors.DictCursor)
        self.islemidd = []
        self.x1_gozlem = []
        self.x2_gozlem = []
        self.x3_gozlem = []
        self.x1 = self.x1_degerler
        self.x2 = self.x2_degerler
        self.x1_min = min(self.x1)
        self.x2_min = min(self.x2)
        self.x1_max = max(self.x1)
        self.x2_max = max(self.x2)
        print("Degerler : {},{},{},{}".format(self.x1_min,self.x1_max,self.x2_min,self.x2_max))
        self.eskix1 = float(self.nokta1.get())
        self.eskix2 = float(self.nokta2.get())
        self.x1_normalize ="{:.2f}".format((self.eskix1 - self.x1_min) /  (self.x1_max - self.x1_min))
        self.x2_normalize ="{:.2f}".format((self.eskix2 - self.x2_min) /  (self.x2_max - self.x2_min))
        print(self.x1_normalize)
        print(self.x2_normalize)
        if(self.sütünn_sayisi == 4):
            self.x3 = self.x3_degerler
            self.x3_max = max(self.x3)
            self.x3_min = min(self.x3)
            self.eskix3 = float(self.nokta3.get())
            self.x3_normalize = "{:.2f}".format((self.eskix3 - self.x3_min) /  (self.x3_max - self.x3_min))

        for i in range(0,len(self.x1)):
            self.x1_gozlem.append("{:.2f}".format( (self.x1[i] - self.x1_min) /  (self.x1_max - self.x1_min)))
            self.x2_gozlem.append("{:.2f}".format( (self.x2[i] - self.x2_min) /  (self.x2_max - self.x2_min)))

            if(self.sütünn_sayisi == 4):
                self.x3_gozlem.append("{:.2f}".format( (self.x3[i] - self.x3_min) /  (self.x3_max - self.x3_min)))
        if(self.sütünn_sayisi == 4):
            print(self.x3_gozlem)
        with connection.cursor() as cursor:
            sorgu = "SELECT islem_id FROM degerler"
            cursor.execute(sorgu)
            for i in cursor.fetchall():
                self.islemidd.append(i["islem_id"])
            for i in range(0,len(self.x1_gozlem)):
                cursor.execute("UPDATE degerler SET x1_gozlem='%f'  WHERE islem_id='%d'"%(float(self.x1_gozlem[i]),self.islemidd[i]))
                cursor.execute("UPDATE degerler SET x2_gozlem='%f'  WHERE islem_id='%d'"%(float(self.x2_gozlem[i]),self.islemidd[i]))
                if(self.sütünn_sayisi == 4):
                    cursor.execute("UPDATE degerler SET x3_gozlem='%f'  WHERE islem_id='%d'"%(float(self.x3_gozlem[i]),self.islemidd[i]))
            self.listekutusu.insert(END,"------------------------")
            self.listekutusu.insert(END,"------------------------")
            self.listekutusu.insert(END,"Gözlem Değerleri")
            if(self.sütünn_sayisi == 4):
                self.listekutusu.insert(END,"{:^10}{:^10}{:^10}{:^10}{:^10}{:^10}".format("X1","X2","X3","X1'","X2'","X3'"))
            else:
                self.listekutusu.insert(END,"{:^10}{:^10}{:^10}{:^10}".format("X1","X2","X1'","X2'"))
            for i in range (0,len(self.x1_gozlem)):
                if(self.sütünn_sayisi == 3):

                     self.listekutusu.insert(END,"{:^10}{:^10}{:^10}{:^10}".format(self.x1[i],self.x2[i],self.x1_gozlem[i] , self.x2_gozlem[i]))


                else:
                    self.listekutusu.insert(END,"{:^10}{:^10}{:^10}{:^10}{:^10}{:^10}".format(self.x1[i],self.x2[i],self.x3[i],self.x1_gozlem[i] , self.x2_gozlem[i],self.x3_gozlem[i]))

        self.x1minkutu.configure(state="normal")
        self.x1minkutu.insert(END,str(self.x1_min))
        self.x1minkutu.configure(state="readonly")

        self.x1maxkutu.configure(state="normal")
        self.x1maxkutu.insert(END,str(self.x1_max))
        self.x1maxkutu.configure(state="readonly")

        self.x2minkutu.configure(state="normal")
        self.x2minkutu.insert(END,str(self.x2_min))
        self.x2minkutu.configure(state="readonly")

        self.x2maxkutu.configure(state="normal")
        self.x2maxkutu.insert(END,str(self.x2_max))
        self.x2maxkutu.configure(state="readonly")

        self.yeninokta1.configure(state="normal")
        self.yeninokta1.insert(END,str(self.x1_normalize))
        self.yeninokta1.configure(state="readonly")

        self.yeninokta2.configure(state="normal")
        self.yeninokta2.insert(END,str(self.x2_normalize))
        self.yeninokta2.configure(state="readonly")


        if(self.sütünn_sayisi==4):
            self.x3minkutu.configure(state="normal")
            self.x3minkutu.insert(END,str(self.x1_min))
            self.x3minkutu.configure(state="readonly")

            self.x3maxkutu.configure(state="normal")
            self.x3maxkutu.insert(END,str(self.x3_max))
            self.x3maxkutu.configure(state="readonly")

            self.yeninokta3.configure(state="normal")
            self.yeninokta3.insert(END,str(self.x3_normalize))
            self.yeninokta3.configure(state="readonly")

    pass

    def GrafikGöster(self):
        try:
            connection = pymysql.connect(host='localhost',
                                         user='root',
                                         password='',
                                         db='verimadenciligi',
                                         charset='utf8mb4',
                                         cursorclass=pymysql.cursors.DictCursor)
            wb = openpyxl.load_workbook(self.filename)
            tablo = wb.get_sheet_by_name('Sheet1')
            tablo = wb.active
            self.sütün_sayisi = tablo.max_column
            self.x1_deger = []
            self.x2_deger = []
            self.x3_deger = []
            with connection.cursor() as cursor:

                            sorgu = "SELECT x1 , x2 , x3 FROM degerler"
                            cursor.execute(sorgu)
                            for i in cursor.fetchall():
                                self.x1_deger.append(i["x1"])
                                self.x2_deger.append(i["x2"])
                                if(self.sütün_sayisi == 4):
                                    self.x3_deger.append(i["x3"])
            print("---------------")

            plt.xlabel("x1")
            plt.ylabel("x2")
            bc2=plt.scatter(self.x1_deger, self.x2_deger)
            plt.grid(True)
            plt.show()
        except:
            messagebox.showerror("HATA","Değer grafiğini görebilmek için değerleri giriniz")


    def NormalHesapla(self):
        #try:
                wb = openpyxl.load_workbook(self.filename)
                tablo = wb.get_sheet_by_name('Sheet1')
                tablo = wb.active
                sütün_sayisi = tablo.max_column
                self.x1deger = []
                self.x2deger = []
                self.x3deger = []
                self.uzaklik = []
                self.uzaklik_degerleri=[]
                connection = pymysql.connect(host='localhost',
                                             user='root',
                                             password='',
                                             db='verimadenciligi',
                                             charset='utf8mb4',
                                             cursorclass=pymysql.cursors.DictCursor)

                #if((self.nokta1.get() != "") and (self.nokta2.get()!="")):
                   # if(self.nokta3.get() != ""):
                self.Nokta1 = float(self.nokta1.get())
                self.Nokta2 = float(self.nokta2.get())
                self.islemid = []
                if(sütün_sayisi == 4):
                    self.Nokta3 = float(self.nokta3.get())
                with connection.cursor() as cursor:

                            sorgu = "SELECT islem_id , x1 , x2 , x3 FROM degerler"
                            cursor.execute(sorgu)
                            for i in cursor.fetchall():
                                self.x1deger.append(i["x1"])
                                self.x2deger.append(i["x2"])
                                self.islemid.append(i["islem_id"])

                                if(sütün_sayisi == 4):
                                    self.x3deger.append(i["x3"])



                uzunluk = (len(self.x1deger))
                for i in range (0,uzunluk):
                    gecici = (self.x1deger[i] - self.Nokta1) * (self.x1deger[i] - self.Nokta1)
                    gecici2 = (self.x2deger[i] - self.Nokta2) * (self.x2deger[i] - self.Nokta2)
                    if(sütün_sayisi == 4):
                        gecici3 = (self.x3deger[i] - self.Nokta3) * (self.x3deger[i] - self.Nokta3)
                    if(sütün_sayisi == 4):
                        karekök3 = math.sqrt(gecici+gecici2+gecici3)
                        karekök4 ="{:.2f}".format(karekök3)
                        self.uzaklik.append(karekök4)
                    else:
                        karekök = math.sqrt(gecici+gecici2)
                        karekök2 = "{:.2f}".format(karekök)
                        self.uzaklik.append(karekök2)
                with connection.cursor() as cursor:
                        for i in range(0,uzunluk):
                            cursor.execute("UPDATE degerler SET uzaklık_deger='%f'  WHERE islem_id='%d'"%(float(self.uzaklik[i]),self.islemid[i]))

                if(sütün_sayisi == 3):

                    messagebox.showinfo("Onay","Uzaklık Hesaplandi")
                    self.listekutusu.insert(END,"Noktaların değere uzaklıkları bulundu...")
                    self.listekutusu.insert(END,"  X1-------X2---------Uzaklık")
                else:
                    messagebox.showinfo("Onay","Uzaklık Hesaplandi")
                    self.listekutusu.insert(END,"Noktaların değere uzaklıkları bulundu...")
                    self.listekutusu.insert(END,"  X1-------X2-----X3---------Uzaklık")


                for i in range (0,(uzunluk)):
                    if(sütün_sayisi == 3):

                        self.listekutusu.insert(END,"{:^10}{:^10}{:^30}".format(self.x1deger[i] , self.x2deger[i] , self.uzaklik[i]))

                    else:
                        self.listekutusu.insert(END,"{:^10}{:^10}{:^10}{:^30}".format(self.x1deger[i] , self.x2deger[i],self.x3deger[i] , self.uzaklik[i]))

                print(self.uzaklik)
        #except:
           # messagebox.showwarning("HATA","İşlem Yapılacak Dosyayı Seciniz ve Degerleri Giriniz")

        #print("{0} \n {1} \n {2}".format(self.x1deger,self.x2deger,self.x3deger) )

       # else:
            #messagebox.showwarning("Eksik Deger","Nokta Bilgilerini Giriniz")
    def NormalizeHesapla(self):
        try:
                connection = pymysql.connect(host='localhost',
                                             user='root',
                                             password='',
                                             db='verimadenciligi',
                                             charset='utf8mb4',
                                             cursorclass=pymysql.cursors.DictCursor)
                self.x1deneme=[]
                with connection.cursor() as cursor:
                    sql = "SELECT islem_id FROM `degerler`"
                    cursor.execute(sql)
                    for abc in cursor.fetchall():
                        self.x1deneme.append(abc["islem_id"])
                wb = openpyxl.load_workbook(self.filename)
                tablo = wb.get_sheet_by_name('Sheet1')
                tablo = wb.active
                sütün_sayisi = tablo.max_column
                self.x1gozlem = []
                self.x2gozlem = []
                self.x3gozlem = []
                self.uzaklikgozlem = []
                self.uzaklikkgozlem = []
                for i in range(0,len(self.x1_gozlem)):
                    self.x1gozlem.append(float(self.x1_gozlem[i]))
                    self.x2gozlem.append(float(self.x2_gozlem[i]))

                    if(sütün_sayisi ==4):
                        self.x3gozlem.append(float(self.x3_gozlem[i]))

                #if((self.nokta1.get() != "") and (self.nokta2.get()!="")):
                   # if(self.nokta3.get() != ""):
                self.nNokta1 = float(self.yeninokta1.get())
                self.nNokta2 = float(self.yeninokta2.get())
                if(sütün_sayisi == 4):
                    self.nNokta3 = float(self.yeninokta3.get())

                uzunluk = (len(self.x1gozlem))
                for i in range (0,uzunluk):
                    gecici = (self.x1gozlem[i] - self.nNokta1) * (self.x1gozlem[i] - self.nNokta1)
                    gecici2 = (self.x2gozlem[i] - self.nNokta2) * (self.x2gozlem[i] - self.nNokta2)
                    if(sütün_sayisi == 4):
                        gecici3 = (self.x3gozlem[i] - self.nNokta3) * (self.x3gozlem[i] - self.nNokta3)
                    if(sütün_sayisi == 4):
                        karekök3 = math.sqrt(gecici+gecici2+gecici3)
                        karekök4 ="{:.2f}".format(karekök3)
                        self.uzaklikgozlem.append(karekök4)
                    else:
                        karekök = math.sqrt(gecici+gecici2)
                        karekök2 = "{:.2f}".format(karekök)
                        self.uzaklikgozlem.append(karekök2)
                try:
                    i=0
                    for i in range(0,uzunluk):
                        print(self.uzaklikgozlem[i])
                        self.uzaklikkgozlem.append(float(self.uzaklikgozlem[i]))


                    y=0
                    with connection.cursor() as cursor:

                            for y in range(0,uzunluk):
                                cursor.execute("UPDATE degerler SET uzaklık_deger='%f'  WHERE islem_id='%d'"%(self.uzaklikkgozlem[y],self.x1deneme[y]))
                except:
                    messagebox.showinfo("hata","asd")

                if(sütün_sayisi == 3):

                    messagebox.showinfo("Onay","Uzaklık Hesaplandi")
                    self.listekutusu.insert(END,"------------------------")
                    self.listekutusu.insert(END,"------------------------")
                    self.listekutusu.insert(END,"Noktaların değere uzaklıkları bulundu...")
                    self.listekutusu.insert(END,"  X1-------X2---------Uzaklık")
                else:
                    messagebox.showinfo("Onay","Uzaklık Hesaplandi")
                    self.listekutusu.insert(END,"------------------------")
                    self.listekutusu.insert(END,"------------------------")
                    self.listekutusu.insert(END,"Noktaların değere uzaklıkları bulundu...")
                    self.listekutusu.insert(END,"  X1-------X2-----X3---------Uzaklık")

                for i in range (0,(uzunluk)):
                    if(sütün_sayisi == 3):

                        self.listekutusu.insert(END,"{:^10}{:^10}{:^30}".format(self.x1gozlem[i] , self.x2gozlem[i] , self.uzaklikgozlem[i]))

                    else:
                        self.listekutusu.insert(END,"{:^10}{:^10}{:^10}{:^30}".format(self.x1gozlem[i] , self.x2gozlem[i],self.x3gozlem[i] , self.uzaklikgozlem[i]))
        except:
            messagebox.showwarning("HATA","İşlem Yapılacak Dosyayı Seciniz ve Degerleri Giriniz")
    def Uzaklik_Hesapla(self):
        try:

            if(self.CheckVar1.get() == 1):
                self.NormalizeHesapla()
            else:
                self.NormalHesapla()
        except:
            messagebox.showerror("Hata","Uzaklik Hesaplanabilmesi için öncelikli işlemleri yapınız")


    def Temizle(self):
        connection = pymysql.connect(host='localhost',
                                     user='root',
                                     password='',
                                     db='verimadenciligi',
                                     charset='utf8mb4',
                                     cursorclass=pymysql.cursors.DictCursor)
        baglanti = connection.cursor()
        sorgu = "SELECT x1 FROM degerler "
        baglanti.execute(sorgu)
        tüm_satirlar = len(baglanti.fetchall())
        with connection.cursor() as cursor:
            for i in range(0, tüm_satirlar ):
                sorgu = "DELETE FROM degerler WHERE islem_id =  '%d'" % (i)
                cursor.execute(sorgu)


    def dosyaSec(self):
        self.Temizle()
        self.Sifirla()
        try:
            self.filename = filedialog.askopenfilename(initialdir="/", title="Select file",
                                                  filetypes=(("xlsx field", "*.xlsx"), ("all files", "*.*")))
            if(self.filename!=""):
                self.listekutusu.insert(0,"Dosya Yolu : ")
                self.listekutusu.insert(1,self.filename)
                self.listekutusu.insert(2,"Degerler Veri Tabanına Aktarılıyor")
                self.listekutusu.insert(3,"Aktarılan Degerler : ")
                self.degerleri_oku(self.filename)
                self.kdeger.configure(state="readonly")
        except:
            messagebox.showerror("HATA","HATALI SECİM YAPTINIZ")




    def degerleri_oku(self,yol):
        connection = pymysql.connect(host='localhost',
                                     user='root',
                                     password='',
                                     db='verimadenciligi',
                                     charset='utf8mb4',
                                     cursorclass=pymysql.cursors.DictCursor)

        wb = openpyxl.load_workbook(yol)
        tablo = wb.get_sheet_by_name('Sheet1')
        tablo = wb.active
        self.sütun_sayisi = tablo.max_column
        satir_sayisi = tablo.max_row

        self.x1_degerler = []
        self.x2_degerler = []
        self.x3_degerler = []
        self.sınıf_degerler = []
        islemid = []
        if (self.sütun_sayisi == 3):
            self.nokta3.configure(state='disabled')
            self.nokta2.configure(state='active')
            self.nokta1.configure(state='active')
            for i in range(0, satir_sayisi - 1):
                islemid.append(i)
            for i in range(2, satir_sayisi + 1):
                deger = tablo.cell(row=i, column=1)
                self.x1_degerler.append(deger.value)
                deger = tablo.cell(row=i, column=2)
                self.x2_degerler.append(deger.value)
                deger = tablo.cell(row=i, column=3)
                self.sınıf_degerler.append(str(deger.value))

            with connection.cursor() as cursor:
                for i in range(0, len(self.x1_degerler)):
                    yazdir = "INSERT INTO degerler(islem_id,x1, x2, x3, durum,uzaklık_deger,agırlık_deger,x1_gozlem,x2_gozlem,x3_gozlem) \
                      VALUES('%d','%f', '%f','%f', '%s','%f','%f','%f','%f','%f' )" % \
                             (islemid[i], self.x1_degerler[i], self.x2_degerler[i], -1, self.sınıf_degerler[i],0.0,0.0,0.0,0.0,0.0)
                    connection.commit()
                    cursor.execute(yazdir)
                    self.listekutusu.insert(END,"{:^10}{:^10}{:^30}".format(self.x1_degerler[i] , self.x2_degerler[i]  , self.sınıf_degerler[i]))


        elif (self.sütun_sayisi == 4):
            self.nokta3.configure(state='active')
            self.nokta2.configure(state='active')
            self.nokta1.configure(state='active')
            for i in range(0, satir_sayisi - 1):
                islemid.append(i)
            for i in range(2, satir_sayisi + 1):
                deger = tablo.cell(row=i, column=1)
                self.x1_degerler.append(deger.value)
                deger = tablo.cell(row=i, column=2)
                self.x2_degerler.append(deger.value)
                deger = tablo.cell(row=i, column=3)
                self.x3_degerler.append(deger.value)
                deger = tablo.cell(row=i, column=4)
                self.sınıf_degerler.append(str(deger.value))

            with connection.cursor() as cursor:
                for i in range(0, len(self.x1_degerler)):
                    yazdir2 = "INSERT INTO degerler(islem_id,x1, x2, x3, durum,uzaklık_deger,agırlık_deger,x1_gozlem,x2_gozlem,x3_gozlem) \
                      VALUES('%d','%f', '%f','%f', '%s','%f','%f','%f','%f','%f' )" % \
                              (islemid[i], self.x1_degerler[i], self.x2_degerler[i], self.x3_degerler[i], self.sınıf_degerler[i],0.0,0.0,0.0,0.0,0.0)
                    connection.commit()
                    cursor.execute(yazdir2)
                    self.listekutusu.insert(END,"{:^10}{:^10}{:^10}{:^30}".format(self.x1_degerler[i] , self.x2_degerler[i] , self.x3_degerler[i] , self.sınıf_degerler[i]))




    def Sifirla(self):
        self.nokta3.configure(state='disabled')
        self.nokta2.configure(state='disabled')
        self.nokta1.configure(state='disabled')
        self.listekutusu.delete(0,END)
        self.kdeger.set("")
        self.nokta1.set("")
        self.nokta2.set("")
        self.nokta3.set("")
        self.CheckVar1.set(0)
        self.CheckVar2.set(0)
        self.CheckVar3.set(0)
        self.x1minkutu.configure(state="normal")
        self.x1minkutu.delete(0,END)
        self.x1minkutu.configure(state="readonly")

        self.x1maxkutu.configure(state="normal")
        self.x1maxkutu.delete(0,END)
        self.x1maxkutu.configure(state="readonly")

        self.x2minkutu.configure(state="normal")
        self.x2minkutu.delete(0,END)
        self.x2minkutu.configure(state="readonly")

        self.x2maxkutu.configure(state="normal")
        self.x2maxkutu.delete(0,END)
        self.x2maxkutu.configure(state="readonly")

        self.x3minkutu.configure(state="normal")
        self.x3minkutu.delete(0,END)
        self.x3minkutu.configure(state="readonly")

        self.x3maxkutu.configure(state="normal")
        self.x3maxkutu.delete(0,END)
        self.x3maxkutu.configure(state="readonly")

        self.yeninokta1.configure(state="normal")
        self.yeninokta1.delete(0,END)
        self.yeninokta1.configure(state="readonly")

        self.yeninokta2.configure(state="normal")
        self.yeninokta2.delete(0,END)
        self.yeninokta2.configure(state="readonly")

        self.yeninokta3.configure(state="normal")
        self.yeninokta3.delete(0,END)
        self.yeninokta3.configure(state="readonly")

        self.kdeger.configure(state="disabled")
        self.normalize.configure(state="disabled")




pencere = Tk()
basla = görsel(pencere)
pencere.colormapwindows()

pencere.mainloop()